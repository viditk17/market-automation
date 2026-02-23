import pandas as pd
import tkinter as tk
from tkinter import filedialog
import io

# Hide the main tkinter window
root = tk.Tk()
root.withdraw()

# Helper function for file selection
def select_file(title, filetypes):
    """Open file dialog and return selected file path"""
    filepath = filedialog.askopenfilename(title=title, filetypes=filetypes)
    if not filepath:
        raise Exception(f"No file selected for: {title}")
    return filepath

print("=" * 60)
print("WORK ASSIGNMENT SUMMARY AUTOMATION - COMPLETE")
print("=" * 60)

# Step 1: Select files
print("\n📁 Step 1: Files select karo...")
print("Pehle BRANCH_MASTER_XL.xlsx select karo")
branch_file = select_file("Select BRANCH_MASTER_XL.xlsx", [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])

print("\nAb 0016_work_assign_summary_xl file select karo")
summary_file = select_file("Select 0016_work_assign_summary_xl", [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])

# ============ NEW: CSV UPLOADS ============
print("\nAb CANCEL REMARK REPORT CSV select karo")
cancel_csv_file = select_file("Select CANCEL REMARK REPORT CSV", [("CSV files", "*.csv"), ("All files", "*.*")])

print("\nAb CHALLENGE PRICE REPORT CSV select karo")
challenge_csv_file = select_file("Select CHALLENGE PRICE REPORT CSV", [("CSV files", "*.csv"), ("All files", "*.*")])

print("\nAb VEHICLE HIRING INCENTIVE CSV select karo")
staff_csv_file = select_file("Select VEHICLE HIRING INCENTIVE CSV", [("CSV files", "*.csv"), ("All files", "*.*")])

# ============ NEW: BROKER MASTER CSV ============
print("\nAb BROKER MASTER CSV file select karo")
broker_csv_file = select_file("Select BROKER MASTER CSV", [("CSV files", "*.csv"), ("All files", "*.*")])
print(f" - Broker Master: {broker_csv_file}")

# ===============================================

# ============ NEW: HO AND BRANCH SEGREGATION ============
print("\nAb HO and branch segregation file select karo")
ho_file = select_file("Select HO and branch segregation file", [("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")])

# ============ NEW: Read HO file and create ho_employee_codes ============
try:
    if str(ho_file).lower().endswith('.csv'):
        ho_df = pd.read_csv(ho_file)
    else:
        ho_df = pd.read_excel(ho_file)
    
    if 'EMPLOYEE CODE' in ho_df.columns:
        ho_employee_codes = (
            ho_df['EMPLOYEE CODE']
            .astype(str)
            .str.strip()
            .replace(['nan', 'None', ''], pd.NA)
            .dropna()
            .unique()
            .tolist()
        )
        print(f" ✅ HO employee codes loaded: {len(ho_employee_codes)}")
    else:
        ho_employee_codes = []
        print(" ⚠️ WARNING: 'EMPLOYEE CODE' column not found in HO file!")
except Exception as e:
    ho_employee_codes = []
    print(f" ⚠️ WARNING: HO file read failed: {e}")

# ========================================================
print(f" - HO File: {ho_file}")
# ========================================================

print(f"\n✅ Files uploaded successfully!")
print(f" - Branch Master: {branch_file}")
print(f" - Summary File: {summary_file}")
print(f" - Cancel Report: {cancel_csv_file}")
print(f" - Challenge Report: {challenge_csv_file}")
print(f" - Staff Detail: {staff_csv_file}")
print(f" - Broker Master: {broker_csv_file}")
print(f" - HO File: {ho_file}")

# Step 2: Read files
print("\n📖 Step 2: Files read kar rahe hain...")
branch_master_df = pd.read_excel(branch_file)

# Pehli sheet ko use karo
excel_file = pd.ExcelFile(summary_file)
sheet_name = excel_file.sheet_names[0]
summary_df = pd.read_excel(summary_file, sheet_name=sheet_name)

print(f" - Branch Master rows: {len(branch_master_df)}")
print(f" - Summary rows: {len(summary_df)}")
print(f" - Sheet name: {sheet_name}")

# CANCEL REMARK REPORT - CORRECT READING (skip title row)
print("\n📊 Step 2.1: CANCEL, CHALLENGE and STAFF reports read kar rahe hain...")

# Read Cancel Remark Report - skip first title row, second row is header
cancel_df_report = pd.read_csv(cancel_csv_file, skiprows=1)
print(f" - Cancel Remark Report: {len(cancel_df_report)} rows x {len(cancel_df_report.columns)} columns")
print(f"   Columns: {list(cancel_df_report.columns)[:5]}...")

# CHALLENGE PRICE REPORT
challenge_df = pd.read_csv(challenge_csv_file, skiprows=1)
print(f" - Challenge Price Report: {len(challenge_df)} rows x {len(challenge_df.columns)} columns")

# STAFF DETAIL REPORT (Vehicle Hiring Incentive)
staff_detail_df = pd.read_csv(staff_csv_file)
print(f" - Staff Detail Report: {len(staff_detail_df)} rows x {len(staff_detail_df.columns)} columns")

# ============ NEW: Step 2.3 - Read Broker Master CSV ============
print("\n📋 Step 2.3: BROKER MASTER CSV read kar rahe hain...")
broker_master_df = pd.read_csv(broker_csv_file)
print(f" - Broker Master: {len(broker_master_df)} rows x {len(broker_master_df.columns)} columns")

# ================================================================

# ============ NEW: Step 2.2 - Process Staff Detail Report ============
print("\n🔧 Step 2.2: Staff Detail Report ko process kar rahe hain...")

# Find HIRING column and delete all columns after it
if 'HIRING' in staff_detail_df.columns:
    hiring_col_index = staff_detail_df.columns.tolist().index('HIRING')
    # Keep only columns up to and including HIRING
    keep_columns = staff_detail_df.columns[:hiring_col_index + 1].tolist()
    staff_detail_df = staff_detail_df[keep_columns].copy()
    print(f" - Columns after HIRING deleted. Remaining columns: {len(staff_detail_df.columns)}")
    print(f" - Columns kept: {list(staff_detail_df.columns)}")
    
    # Add ERP and APP columns (initially empty)
    staff_detail_df['ERP'] = 0
    staff_detail_df['APP'] = 0
    print(f" - Added ERP and APP columns")
else:
    print(" ⚠️ WARNING: HIRING column not found in staff detail report!")
    staff_detail_df['ERP'] = 0
    staff_detail_df['APP'] = 0

# =====================================================================

# ===========================================================

# Step 3: Create VLOOKUP mapping (Zone lookup dictionary)
print("\n🔍 Step 3: VLOOKUP setup kar rahe hain...")
zone_lookup = dict(zip(
    branch_master_df['BRANCH_BRANCH_CODE'],
    branch_master_df['ZONE_ZONE_NAME']
))
print(f" - Total zones mapped: {len(zone_lookup)}")

# Step 4: Find column positions for FRM and T_O
print("\n📍 Step 4: FRM aur T_O columns ki position dhundh rahe hain...")
columns_list = summary_df.columns.tolist()

# Find FRM column index
if 'FRM' in columns_list:
    frm_index = columns_list.index('FRM')
    print(f" - FRM column found at position: {frm_index}")
else:
    print(" ⚠️ WARNING: FRM column not found!")
    frm_index = None

# Find T_O column index
if 'T_O' in columns_list:
    to_index = columns_list.index('T_O')
    print(f" - T_O column found at position: {to_index}")
else:
    print(" ⚠️ WARNING: T_O column not found!")
    to_index = None

# Step 5: Add FRM ZONE and TO ZONE columns at specific positions
print("\n➕ Step 5: FRM ZONE aur TO ZONE columns add kar rahe hain...")
frm_zone_data = summary_df['FRM'].map(zone_lookup)
to_zone_data = summary_df['T_O'].map(zone_lookup)

# Insert FRM ZONE right after FRM column
if frm_index is not None:
    summary_df.insert(frm_index + 1, 'FRM ZONE', frm_zone_data)
    print(f" - FRM ZONE column added at position: {frm_index + 1}")
    
    # Update to_index because we added a column before it
    if to_index is not None and to_index > frm_index:
        to_index += 1

# Insert TO ZONE right after T_O column
if to_index is not None:
    summary_df.insert(to_index + 1, 'TO ZONE', to_zone_data)
    print(f" - TO ZONE column added at position: {to_index + 1}")

frm_zone_filled = summary_df['FRM ZONE'].notna().sum()
to_zone_filled = summary_df['TO ZONE'].notna().sum()
print(f" - FRM ZONE filled: {frm_zone_filled}/{len(summary_df)}")
print(f" - TO ZONE filled: {to_zone_filled}/{len(summary_df)}")

# Step 6: Sort by HIRING_NO (ascending)
print("\n📊 Step 6: HIRING_NO se sort kar rahe hain...")
summary_df['HIRING_NO_SORT'] = summary_df['HIRING_NO'].astype(str)
summary_df_sorted = summary_df.sort_values(by='HIRING_NO_SORT', ascending=True)
summary_df_sorted = summary_df_sorted.drop(columns=['HIRING_NO_SORT'])
print(f" ✅ Data sorted successfully!")

# Step 7: Filter rows where HIRING_NO is NOT blank
print("\n🎯 Step 7: HIRING_NO wale rows filter kar rahe hain...")
all_hiring_df = summary_df_sorted[summary_df_sorted['HIRING_NO'].notna()].copy()
print(f" - Total hiring rows: {len(all_hiring_df)} (from {len(summary_df_sorted)} total rows)")

# Step 8: Create "erp" sheet (ENTER_USING = 'ERP')
print("\n🔧 Step 8: ERP sheet bana rahe hain...")
erp_df = summary_df_sorted[summary_df_sorted['ENTER_USING'] == 'ERP'].copy()
print(f" - ERP rows: {len(erp_df)}")

# Step 9: Create "app" sheet (ENTER_USING = 'APP' OR 'FIX RATE')
print("\n📱 Step 9: APP sheet bana rahe hain...")
app_df = summary_df_sorted[summary_df_sorted['ENTER_USING'].isin(['APP', 'FIX RATE'])].copy()
print(f" - APP rows (APP + FIX RATE): {len(app_df)}")

# Step 10: Create "fix" sheet (ENTER_USING = 'FIX RATE')
print("\n🔩 Step 10: FIX sheet bana rahe hain...")
fix_df = summary_df_sorted[summary_df_sorted['ENTER_USING'] == 'FIX RATE'].copy()
print(f" - FIX RATE rows: {len(fix_df)}")

# Step 11: Create "att" sheet (LORRY_REQUIRE_STATUS = 'ATT' OR 'IDLE')
print("\n🚛 Step 11: ATT sheet bana rahe hain...")
att_df = summary_df_sorted[summary_df_sorted['LORRY_REQUIRE_STATUS'].isin(['ATT', 'IDLE'])].copy()
print(f" - ATT rows (ATT + IDLE): {len(att_df)}")

# Step 12: Create "cancel" sheet (LORRY_REQUIRE_STATUS = 'REQ. CAN')
print("\n❌ Step 12: CANCEL sheet bana rahe hain...")
cancel_df = summary_df_sorted[summary_df_sorted['LORRY_REQUIRE_STATUS'] == 'REQ. CAN'].copy()
print(f" - CANCEL rows: {len(cancel_df)}")

# Step 13: Create "fail" sheet (LORRY_REQUIRE_STATUS = 'FAIL')
print("\n⚠️ Step 13: FAIL sheet bana rahe hain...")
fail_df = summary_df_sorted[summary_df_sorted['LORRY_REQUIRE_STATUS'] == 'FAIL'].copy()
print(f" - FAIL rows: {len(fail_df)}")

# ============ NEW: Step 13.1 - Process ERP/APP for Staff Detail COUNTIF ============
print("\n🔢 Step 13.1: ERP and APP COUNTIF formulas apply kar rahe hain...")

# Process ERP sheet - Extract HIRING_BY codes
if 'HIRING_BY' in erp_df.columns:
    # Split by "-" and take first part (employee code)
    erp_codes = erp_df['HIRING_BY'].astype(str).str.split('-').str[0].str.strip()
    erp_code_counts = erp_codes.value_counts()
    print(f" - ERP codes extracted: {len(erp_code_counts)} unique codes")
    
    # Apply COUNTIF to staff_detail_df
    if 'EMPLOYEE CODE' in staff_detail_df.columns:
        # Convert EMPLOYEE CODE to string for matching
        staff_detail_df['EMPLOYEE CODE'] = staff_detail_df['EMPLOYEE CODE'].astype(str).str.strip()
        staff_detail_df['ERP'] = staff_detail_df['EMPLOYEE CODE'].map(erp_code_counts).fillna(0).astype(int)
        print(f" ✅ ERP column filled with COUNTIF values (Total: {staff_detail_df['ERP'].sum()})")
    else:
        print(" ⚠️ WARNING: EMPLOYEE CODE column not found in staff detail report!")
else:
    print(" ⚠️ WARNING: HIRING_BY column not found in ERP sheet!")

# Process APP sheet - Extract HIRING_BY codes
if 'HIRING_BY' in app_df.columns:
    # Split by "-" and take first part (employee code)
    app_codes = app_df['HIRING_BY'].astype(str).str.split('-').str[0].str.strip()
    app_code_counts = app_codes.value_counts()
    print(f" - APP codes extracted: {len(app_code_counts)} unique codes")
    
    # Apply COUNTIF to staff_detail_df
    if 'EMPLOYEE CODE' in staff_detail_df.columns:
        staff_detail_df['APP'] = staff_detail_df['EMPLOYEE CODE'].map(app_code_counts).fillna(0).astype(int)
        print(f" ✅ APP column filled with COUNTIF values (Total: {staff_detail_df['APP'].sum()})")
    else:
        print(" ⚠️ WARNING: EMPLOYEE CODE column not found in staff detail report!")
else:
    print(" ⚠️ WARNING: HIRING_BY column not found in APP sheet!")

print(f" 📊 Staff Detail Summary:")
print(f" - Employees with ERP hiring: {(staff_detail_df['ERP'] > 0).sum()}")
print(f" - Employees with APP hiring: {(staff_detail_df['APP'] > 0).sum()}")

# ===================================================================================

# Step 14: Read Query result sheet and add FRM ZONE, TO ZONE
print("\n🔍 Step 14: Query Result sheet se data nikal rahe hain...")

try:
    if 'Query result' in excel_file.sheet_names:
        query_result_df = pd.read_excel(summary_file, sheet_name='Query result')
        print(f" - Query Result rows: {len(query_result_df)}")
        
        query_result_df['FRM ZONE'] = query_result_df['FRM'].map(zone_lookup)
        query_result_df['TO ZONE'] = query_result_df['T_O'].map(zone_lookup)
        print(f" - FRM ZONE and TO ZONE added to Query result")
        
        # ============ NEW: Step 14.1 - Filter Evening Requests ============
        print("\n🌆 Step 14.1: Evening Request filter kar rahe hain (time >= 18:00:00)...")
        
        if 'REQ_DATE' in query_result_df.columns:
            # Convert REQ_DATE to datetime if not already
            query_result_df['REQ_DATE'] = pd.to_datetime(query_result_df['REQ_DATE'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
            
            # Extract time component
            query_result_df['REQ_TIME'] = query_result_df['REQ_DATE'].dt.time
            
            # Filter rows where time >= 18:00:00
            import datetime
            evening_cutoff = datetime.time(18, 0, 0)
            evening_request_df = query_result_df[query_result_df['REQ_TIME'] >= evening_cutoff].copy()
            
            # Remove temporary time column
            evening_request_df = evening_request_df.drop(columns=['REQ_TIME'])
            query_result_df = query_result_df.drop(columns=['REQ_TIME'])
            
            print(f" ✅ Evening requests filtered: {len(evening_request_df)} rows (from {len(query_result_df)} total)")
        else:
            print(" ⚠️ WARNING: REQ_DATE column not found in Query result!")
            evening_request_df = pd.DataFrame()  # Empty dataframe
        
        # ==================================================================
        
        if 'LORRY_RATE_AMOUNT' in query_result_df.columns:
            temp_df = query_result_df[['FRM ZONE', 'TO ZONE', 'LORRY_RATE_AMOUNT']].copy()
            print(f" - Temp sheet created with 3 columns")
            
            # Step 15: Text-to-Column on LORRY_RATE_AMOUNT
            print("\n✂️ Step 15: Text-to-Column kar rahe hain LORRY_RATE_AMOUNT pe...")
            rate_split = temp_df['LORRY_RATE_AMOUNT'].astype(str).str.split(',', expand=True)
            max_columns = rate_split.shape[1]
            print(f" - Maximum RATE columns: {max_columns}")
            
            rate_columns = {i: f'RATE{i+1}' for i in range(max_columns)}
            rate_split = rate_split.rename(columns=rate_columns)
            
            temp_df = temp_df.drop(columns=['LORRY_RATE_AMOUNT'])
            temp_df = pd.concat([temp_df.reset_index(drop=True), rate_split.reset_index(drop=True)], axis=1)
            
            print(f" ✅ Text-to-Column completed!")
            
            # Step 16: Clean RATE columns
            print("\n🧹 Step 16: RATE columns clean kar rahe hain (keeping last 2 parts + cleaning vendor name)...")
            
            def clean_rate_value(rate_value):
                if pd.isna(rate_value):
                    return rate_value
                
                rate_str = str(rate_value).strip()
                if rate_str.lower() in ['nan', 'none', '']:
                    return rate_value
                
                parts = rate_str.split('-')
                non_empty_parts = [p.strip() for p in parts if p.strip()]
                
                if len(non_empty_parts) >= 2:
                    vendor_code = non_empty_parts[-2]
                    vendor_name = ' '.join(non_empty_parts[-1].split())
                    return f"{vendor_code}-{vendor_name}"
                elif len(non_empty_parts) == 1:
                    return non_empty_parts[0]
                else:
                    return rate_str
            
            rate_col_names = [col for col in temp_df.columns if col.startswith('RATE')]
            for rate_col in rate_col_names:
                temp_df[rate_col] = temp_df[rate_col].apply(clean_rate_value)
                print(f" - {rate_col} cleaned")
            
            print(f" ✅ All RATE columns cleaned!")
            
            # Step 17: Create individual rate sheets
            print("\n📋 Step 17: Individual RATE sheets bana rahe hain...")
            rate_sheets = {}
            for rate_col in rate_col_names:
                rate_num = rate_col.replace('RATE', '')
                sheet_name_rate = f'rate{rate_num}'
                
                rate_df = temp_df[['FRM ZONE', 'TO ZONE', rate_col]].copy()
                rate_sheets[sheet_name_rate] = rate_df
                print(f" - {sheet_name_rate}: {len(rate_df)} rows (before blank removal)")
            
            # Step 18: Remove blank rows from each rate sheet
            print("\n🗑️ Step 18: Blank rows remove kar rahe hain har rate sheet se...")
            for sheet_name_rate, rate_df in rate_sheets.items():
                rate_col = [col for col in rate_df.columns if col.startswith('RATE')][0]
                
                before_count = len(rate_df)
                rate_df_cleaned = rate_df[
                    rate_df[rate_col].notna() &
                    (rate_df[rate_col].astype(str).str.strip() != '') &
                    (rate_df[rate_col].astype(str).str.lower() != 'nan')
                ].copy()
                
                rate_sheets[sheet_name_rate] = rate_df_cleaned
                after_count = len(rate_df_cleaned)
                print(f" - {sheet_name_rate}: {before_count - after_count} blank rows removed ({before_count} → {after_count})")
            
            # Step 19: Merge all rate sheets into rate1
            print("\n🔗 Step 19: Sabhi rate sheets ko rate1 mein merge kar rahe hain...")
            if 'rate1' in rate_sheets:
                combined_rate1 = rate_sheets['rate1'].copy()
                rate1_col = [col for col in combined_rate1.columns if col.startswith('RATE')][0]
                combined_rate1 = combined_rate1.rename(columns={rate1_col: 'RATE1'})
                print(f" - Starting with rate1: {len(combined_rate1)} rows")
                
                for sheet_name_rate in sorted(rate_sheets.keys()):
                    if sheet_name_rate != 'rate1':
                        rate_df = rate_sheets[sheet_name_rate].copy()
                        rate_col = [col for col in rate_df.columns if col.startswith('RATE')][0]
                        rate_df = rate_df.rename(columns={rate_col: 'RATE1'})
                        
                        combined_rate1 = pd.concat([combined_rate1, rate_df], ignore_index=True)
                        print(f" - Appended {sheet_name_rate}: {len(rate_df)} rows")
                
                print(f" ✅ Merged all rate sheets! Total rows in rate1: {len(combined_rate1)}")
                
                # Step 20: Split RATE1 into vendor code and vendor name
                print("\n✂️ Step 20: RATE1 ko split kar rahe hain (vendor code + vendor name)...")
                
                rate1_split = combined_rate1['RATE1'].astype(str).str.split('-', n=1, expand=True)
                combined_rate1['vendor code'] = rate1_split[0] if rate1_split.shape[1] > 0 else ''
                combined_rate1['vendor name'] = rate1_split[1] if rate1_split.shape[1] > 1 else ''
                combined_rate1 = combined_rate1.drop(columns=['RATE1'])
                
                # Normalize
                combined_rate1['vendor code'] = (
                    combined_rate1['vendor code']
                    .astype(str).str.strip()
                    .str.replace(r'\.0$', '', regex=True)
                )
                combined_rate1['vendor name'] = combined_rate1['vendor name'].astype(str).str.strip()
                
                # ✅ Exception fix (BHOLA + LOCAL OWNER VENDOR)
                print("\n🛠️ Exception fix: BHOLA + LOCAL OWNER VENDOR ...")
                
                vcode = combined_rate1['vendor code'].astype(str).str.strip()
                vname = combined_rate1['vendor name'].astype(str).str.strip()
                
                mask_bhola = (
                    vcode.str.contains('BHOLA', case=False, na=False) &
                    vname.str.contains('LOCAL OWNER VENDOR', case=False, na=False)
                )
                
                fixed_count = int(mask_bhola.sum())
                if fixed_count > 0:
                    combined_rate1.loc[mask_bhola, 'vendor name'] = 'BHOLA LOCAL OWNER VENDOR'
                    combined_rate1.loc[mask_bhola, 'vendor code'] = '3150000003'
                    print(f" ✅ Fixed rows: {fixed_count}")
                else:
                    print(" - No matching rows found for exception.")
                
                # ✅ Remove rows where vendor name has any digit
                print("\n🗑️ Removing rows: vendor name me number aa raha hai...")
                
                before_vendorname_clean = len(combined_rate1)
                mask_vendorname_has_number = combined_rate1['vendor name'].astype(str).str.contains(r'\d', na=False)
                removed_vendorname_count = int(mask_vendorname_has_number.sum())
                combined_rate1 = combined_rate1[~mask_vendorname_has_number].copy()
                
                print(f" ✅ Removed {removed_vendorname_count} rows ({before_vendorname_clean} → {len(combined_rate1)})")
                
                # ✅ Remove rows where vendor code is numeric and 7 digits or less
                print("\n🗑️ Removing rows: vendor code sirf number ho aur 7 digits ya kam ho...")
                
                before_vendorcode_clean = len(combined_rate1)
                vcode_clean = (
                    combined_rate1['vendor code']
                    .astype(str).str.strip()
                    .str.replace(r'\.0$', '', regex=True)
                )
                
                mask_vendorcode_short_numeric = vcode_clean.str.fullmatch(r'\d{1,7}', na=False)
                removed_vendorcode_count = int(mask_vendorcode_short_numeric.sum())
                combined_rate1 = combined_rate1[~mask_vendorcode_short_numeric].copy()
                
                print(f" ✅ Removed {removed_vendorcode_count} rows ({before_vendorcode_clean} → {len(combined_rate1)})")
                
                # Step 21: Sort by vendor code
                print("\n📊 Step 21: Vendor code se sort kar rahe hain...")
                combined_rate1 = combined_rate1.sort_values(by='vendor code', ascending=True).copy()
                combined_rate1.reset_index(drop=True, inplace=True)
                print(f" ✅ Sorted by vendor code!")
                
                # Step 22: Count column
                print("\n🧾 Step 22: 'Count' column (values) bana rahe hain + sort (FALSE upar, TRUE neeche) + 1 gap row...")
                
                combined_rate1['Count'] = combined_rate1['vendor code'].astype(str).str.strip().eq(
                    combined_rate1['vendor code'].astype(str).str.strip().shift(-1)
                )
                
                combined_rate1_sorted_count = combined_rate1.sort_values(by='Count', ascending=True, kind='mergesort').copy()
                combined_rate1_sorted_count.reset_index(drop=True, inplace=True)
                
                false_block = combined_rate1_sorted_count[combined_rate1_sorted_count['Count'] == False].copy()
                true_block = combined_rate1_sorted_count[combined_rate1_sorted_count['Count'] == True].copy()
                
                if len(false_block) > 0 and len(true_block) > 0:
                    blank_row = {col: None for col in combined_rate1_sorted_count.columns}
                    gap_df = pd.DataFrame([blank_row])
                    combined_rate1_gap = pd.concat([false_block, gap_df, true_block], ignore_index=True)
                    print(" ✅ Gap row inserted between FALSE and TRUE blocks")
                else:
                    combined_rate1_gap = combined_rate1_sorted_count
                    print(" - Gap row not inserted (either TRUE block or FALSE block missing)")
                
                # Step 23: Count IF column
                print("\n🧮 Step 23: 'Count IF' column bana rahe hain (sirf FALSE rows) + values + sort (largest->smallest)...")
                
                vc_series = combined_rate1_gap['vendor code'].dropna().astype(str).str.strip()
                vc_counts = vc_series.value_counts()
                
                combined_rate1_gap['Count IF'] = None
                mask_false_rows = (combined_rate1_gap['Count'] == False) & (combined_rate1_gap['vendor code'].notna())
                
                combined_rate1_gap.loc[mask_false_rows, 'Count IF'] = (
                    combined_rate1_gap.loc[mask_false_rows, 'vendor code']
                    .astype(str).str.strip()
                    .map(vc_counts)
                )
                
                false_block2 = combined_rate1_gap[combined_rate1_gap['Count'] == False].copy()
                gap_block = combined_rate1_gap[combined_rate1_gap['Count'].isna()].copy()
                true_block2 = combined_rate1_gap[combined_rate1_gap['Count'] == True].copy()
                
                false_block2['Count IF'] = pd.to_numeric(false_block2['Count IF'], errors='coerce')
                false_block2_sorted = false_block2.sort_values(by='Count IF', ascending=False, kind='mergesort').copy()
                false_block2_sorted.reset_index(drop=True, inplace=True)
                
                final_rate1 = pd.concat([false_block2_sorted, gap_block, true_block2], ignore_index=True)
                
                print(f" ✅ Count IF added + FALSE block sorted by Count IF (desc). Final rows: {len(final_rate1)}")
                
                # Step 23.5: Active Vendor sheet
                print("\n✅ Step 23.5: 'active vendor' sheet bana rahe hain (sirf Count == FALSE rows)...")
                active_vendor_df = final_rate1[final_rate1['Count'] == False].copy()
                active_vendor_df.reset_index(drop=True, inplace=True)
                print(f" - active vendor rows: {len(active_vendor_df)}")
                
                # Step 24: Add ERP and APP columns to active vendor
                print("\n📊 Step 24: Active vendor mein ERP aur APP columns add kar rahe hain...")
                
                if 'HIRING_VENDOR_CODE' in erp_df.columns and 'HIRING_VENDOR_CODE' in app_df.columns:
                    def normalize_vendor_code(code):
                        code_str = str(code).strip()
                        if code_str.endswith('.0'):
                            code_str = code_str[:-2]
                        return code_str
                    
                    erp_df['HIRING_VENDOR_CODE_NORMALIZED'] = erp_df['HIRING_VENDOR_CODE'].apply(normalize_vendor_code)
                    app_df['HIRING_VENDOR_CODE_NORMALIZED'] = app_df['HIRING_VENDOR_CODE'].apply(normalize_vendor_code)
                    active_vendor_df['vendor_code_normalized'] = active_vendor_df['vendor code'].apply(normalize_vendor_code)
                    
                    erp_counts = erp_df['HIRING_VENDOR_CODE_NORMALIZED'].value_counts()
                    app_counts = app_df['HIRING_VENDOR_CODE_NORMALIZED'].value_counts()
                    
                    active_vendor_df['ERP'] = active_vendor_df['vendor_code_normalized'].map(erp_counts).fillna(0).astype(int)
                    active_vendor_df['APP'] = active_vendor_df['vendor_code_normalized'].map(app_counts).fillna(0).astype(int)
                    
                    active_vendor_df = active_vendor_df.drop(columns=['vendor_code_normalized'])
                    
                    print(f" ✅ ERP column added (Total count: {active_vendor_df['ERP'].sum()})")
                    print(f" ✅ APP column added (Total count: {active_vendor_df['APP'].sum()})")
                    
                    matched_erp = (active_vendor_df['ERP'] > 0).sum()
                    matched_app = (active_vendor_df['APP'] > 0).sum()
                    print(f" 📊 Vendors with ERP matches: {matched_erp}/{len(active_vendor_df)}")
                    print(f" 📊 Vendors with APP matches: {matched_app}/{len(active_vendor_df)}")
                else:
                    print(" ⚠️ WARNING: HIRING_VENDOR_CODE column not found in erp or app sheets!")
                    active_vendor_df['ERP'] = 0
                    active_vendor_df['APP'] = 0
                
                rate_sheets_final = {'rate1': final_rate1}
                temp_sheet_exists = True
            else:
                print(" ⚠️ WARNING: rate1 sheet not found!")
                rate_sheets_final = {}
                temp_sheet_exists = True
                active_vendor_df = None
        else:
            print(f" ⚠️ WARNING: LORRY_RATE_AMOUNT column not found in Query result")
            temp_sheet_exists = False
            temp_df = None
            rate_sheets_final = {}
            active_vendor_df = None
    else:
        print(" ⚠️ WARNING: 'Query result' sheet not found!")
        temp_sheet_exists = False
        temp_df = None
        rate_sheets_final = {}
        active_vendor_df = None
        evening_request_df = pd.DataFrame()

except Exception as e:
    print(f" ⚠️ ERROR reading query result sheet: {e}")
    import traceback
    traceback.print_exc()
    
    temp_sheet_exists = False
    temp_df = None
    rate_sheets_final = {}
    active_vendor_df = None
    evening_request_df = pd.DataFrame()

# ============ NEW: Step 24.5 - Process Broker Master & Create New Vendor Sheet ============
print("\n🆕 Step 24.5: Broker Master ko process kar rahe hain...")

# Check if Query result exists to get target month/year
if 'query_result_df' in locals() and query_result_df is not None and len(query_result_df) > 0:
    if 'REQ_DATE' in query_result_df.columns:
        # Parse REQ_DATE to extract month and year
        query_result_df['REQ_DATE_PARSED'] = pd.to_datetime(
            query_result_df['REQ_DATE'],
            format='%d-%m-%Y %H:%M:%S',
            errors='coerce'
        )
        
        # Get the first non-null date to determine target month/year
        first_date = query_result_df['REQ_DATE_PARSED'].dropna().iloc[0] if len(query_result_df['REQ_DATE_PARSED'].dropna()) > 0 else None
        
        if first_date is not None:
            target_month = first_date.month
            target_year = first_date.year
            print(f" - Target Month/Year detected: {target_month}/{target_year}")
            
            # Now filter broker_master_df by ENTER_DATE
            if 'ENTER_DATE' in broker_master_df.columns:
                # DEBUG: Show sample ENTER_DATE values
                print(f" 📊 Sample ENTER_DATE values (first 10):")
                print(broker_master_df['ENTER_DATE'].head(10).to_string())
                
                # Parse ENTER_DATE with FLEXIBLE format ✅ FIX APPLIED
                # Handles: "1/12/2025", "01/12/2025", "1/12/2025, 12:36 PM"
                broker_master_df['ENTER_DATE_PARSED'] = pd.to_datetime(
                    broker_master_df['ENTER_DATE'],
                    dayfirst=True,  # ✅ Interpret as DD/MM/YYYY (not MM/DD/YYYY)
                    errors='coerce'
                )
                
                # DEBUG: Show parsed dates
                print(f" 📅 Parsed dates (first 10):")
                print(broker_master_df['ENTER_DATE_PARSED'].head(10).to_string())
                
                # Count successfully parsed dates
                valid_dates = broker_master_df['ENTER_DATE_PARSED'].notna().sum()
                print(f" ✅ Successfully parsed: {valid_dates}/{len(broker_master_df)}")
                
                # Filter for target month and year
                new_vendor_df = broker_master_df[
                    (broker_master_df['ENTER_DATE_PARSED'].dt.month == target_month) &
                    (broker_master_df['ENTER_DATE_PARSED'].dt.year == target_year)
                ].copy()
                
                # Drop temporary parsed column
                new_vendor_df = new_vendor_df.drop(columns=['ENTER_DATE_PARSED'])
                
                print(f" 🎯 Filtered rows for {target_month}/{target_year}: {len(new_vendor_df)}")
                
                if len(new_vendor_df) > 0:
                    print(f" ✅ New vendor data ready!")
                else:
                    print(f" ⚠️ No vendors found for {target_month}/{target_year}")
            else:
                print(" ⚠️ WARNING: ENTER_DATE column not found in Broker Master!")
                print(f"   Available columns: {broker_master_df.columns.tolist()}")
                new_vendor_df = pd.DataFrame()
        else:
            print(" ⚠️ WARNING: No valid dates found in REQ_DATE column!")
            new_vendor_df = pd.DataFrame()
    else:
        print(" ⚠️ WARNING: REQ_DATE column not found in Query result!")
        new_vendor_df = pd.DataFrame()
else:
    print(" ⚠️ WARNING: Query result sheet not found or empty!")
    new_vendor_df = pd.DataFrame()

# ============ NEW: Step 26 - Read all sheets from summary file for SUMMARY sheet ============
print("\n📊 Step 26: Summary file ke sheets read kar rahe hain for SUMMARY sheet calculations...")

# Initialize variables for SUMMARY sheet
query_result_for_summary = None
all_hiring_for_summary = None
att_for_summary = None
cancel_for_summary = None
fail_for_summary = None
fix_for_summary = None
app_for_summary = None
erp_for_summary = None

# Check and read each sheet (use already loaded dataframes where possible)
for sheet in excel_file.sheet_names:
    sheet_lower = sheet.lower()
    if 'query' in sheet_lower and 'result' in sheet_lower:
        if 'query_result_df' in locals():
            query_result_for_summary = query_result_df
        else:
            query_result_for_summary = pd.read_excel(summary_file, sheet_name=sheet)
        print(f" ✅ Found 'Query result' sheet: {len(query_result_for_summary)} rows")

# Use already created dataframes for other sheets
all_hiring_for_summary = all_hiring_df
att_for_summary = att_df
cancel_for_summary = cancel_df
fail_for_summary = fail_df
fix_for_summary = fix_df
app_for_summary = app_df
erp_for_summary = erp_df

print(f" ✅ Using existing dataframes:")
print(f"   - all hiring: {len(all_hiring_for_summary)} rows")
print(f"   - att: {len(att_for_summary)} rows")
print(f"   - cancel: {len(cancel_for_summary)} rows")
print(f"   - fail: {len(fail_for_summary)} rows")
print(f"   - fix: {len(fix_for_summary)} rows")
print(f"   - app: {len(app_for_summary)} rows")
print(f"   - erp: {len(erp_for_summary)} rows")

# ===================================================================================

# ============ NEW: Step 27 - Create BEAUTIFUL SUMMARY sheet ============
print("\n📋 Step 27: SUMMARY sheet create kar rahe hain (ENHANCED FORMATTING)...")

# Calculate all values (COUNT - 1 as per requirement)
total_generate_req = len(query_result_for_summary) if query_result_for_summary is not None else 0
total_placed_vehicle = len(all_hiring_for_summary) if all_hiring_for_summary is not None else 0
total_att = len(att_for_summary) if att_for_summary is not None else 0
cancel_req = len(cancel_for_summary) if cancel_for_summary is not None else 0
failure = len(fail_for_summary) if fail_for_summary is not None else 0

total_hiring = len(all_hiring_for_summary) if all_hiring_for_summary is not None else 0
fix_rate = len(fix_for_summary) if fix_for_summary is not None else 0
app_hiring = len(app_for_summary) if app_for_summary is not None else 0
erp_hiring = len(erp_for_summary) if erp_for_summary is not None else 0

# Calculate percentages
total_hiring_pct = (total_hiring / total_generate_req) if total_generate_req > 0 else 0
fix_rate_pct = (fix_rate / total_hiring) if total_hiring > 0 else 0
app_hiring_pct = (app_hiring / total_hiring) if total_hiring > 0 else 0
erp_hiring_pct = (erp_hiring / total_hiring) if total_hiring > 0 else 0

# Calculate TOTAL ACTIVES
total_actives = len(active_vendor_df) if active_vendor_df is not None else 0
print(f" - TOTAL ACTIVES (active vendor count): {total_actives}")

# Calculate CHALLENGE metrics
if 'REQUEST_NO' in challenge_df.columns:
    total_challenging = len(challenge_df)
else:
    total_challenging = 0

if 'CHALLENGE_FLAG' in challenge_df.columns:
    accepted = (challenge_df['CHALLENGE_FLAG'] == 'Accept').sum()
    accepted_rows = challenge_df[challenge_df['CHALLENGE_FLAG'] == 'Accept']
    
    if 'VENDOR_BID' in accepted_rows.columns and 'CHALLENGE_PRICE' in accepted_rows.columns:
        vendor_bid_sum = pd.to_numeric(accepted_rows['VENDOR_BID'], errors='coerce').fillna(0).sum()
        challenge_price_sum = pd.to_numeric(accepted_rows['CHALLENGE_PRICE'], errors='coerce').fillna(0).sum()
        saving = vendor_bid_sum - challenge_price_sum
    else:
        saving = 0
else:
    accepted = 0
    saving = 0

# Calculate LOAD TYPE metrics
if 'all_hiring_df' in locals() and len(all_hiring_df) > 0:
    if 'LOAD_TYPE' in all_hiring_df.columns:
        ftl_count = (all_hiring_df['LOAD_TYPE'] == 'FTL').sum()
        ftl_part_count = (all_hiring_df['LOAD_TYPE'] == 'FTL+Part Load').sum()
        part_load_count = (all_hiring_df['LOAD_TYPE'] == 'Part Load').sum()
        load_type_subtotal = ftl_count + ftl_part_count + part_load_count
    else:
        ftl_count = ftl_part_count = part_load_count = load_type_subtotal = 0
else:
    ftl_count = ftl_part_count = part_load_count = load_type_subtotal = 0

# Calculate L2-L3 VEHICLE PLACED metrics
if 'all_hiring_df' in locals() and len(all_hiring_df) > 0:
    if 'RNK' in all_hiring_df.columns:
        l2_count = (all_hiring_df['RNK'] == 'L2').sum()
        l3_count = (all_hiring_df['RNK'] == 'L3').sum()
        l4_count = (all_hiring_df['RNK'] == 'L4').sum()
        l5_count = (all_hiring_df['RNK'] == 'L5').sum()
        vehicle_placed_subtotal = l2_count + l3_count + l4_count + l5_count
    else:
        l2_count = l3_count = l4_count = l5_count = vehicle_placed_subtotal = 0
else:
    l2_count = l3_count = l4_count = l5_count = vehicle_placed_subtotal = 0

# Create BEAUTIFUL summary structure with proper tables
summary_data = []

# TABLE 1: DEC - 2025 (Main Metrics)
summary_data.append(['DEC - 2025', '', ''])
summary_data.append(['TOTAL GENERATE REQ', total_generate_req, ''])
summary_data.append(['TOTAL PLACED VEHICLE', total_placed_vehicle, ''])
summary_data.append(['TOTAL ATT', total_att, ''])
summary_data.append(['CANCEL REQ', cancel_req, ''])
summary_data.append(['FAILURE', failure, ''])
summary_data.append(['', '', ''])  # Gap row
summary_data.append(['', '', ''])  # Gap row

# TABLE 2: HIRING SUMMARY
summary_data.append(['HIRING SUMMARY', '', ''])
summary_data.append(['TOTAL HIRING', total_hiring, f'{total_hiring_pct:.1%}'])
summary_data.append(['FIX RATE', fix_rate, f'{fix_rate_pct:.1%}'])
summary_data.append(['APP HIRING', app_hiring, f'{app_hiring_pct:.1%}'])
summary_data.append(['ERP HIRING', erp_hiring, f'{erp_hiring_pct:.1%}'])
summary_data.append(['', '', ''])  # Gap row
summary_data.append(['', '', ''])  # Gap row

# TABLE 3: LOAD TYPE
summary_data.append(['LOAD TYPE', '', ''])
summary_data.append(['FTL', ftl_count, ''])
summary_data.append(['FTL+Part Load', ftl_part_count, ''])
summary_data.append(['Part Load', part_load_count, ''])
summary_data.append(['SUB TOTAL', load_type_subtotal, ''])
summary_data.append(['', '', ''])  # Gap row
summary_data.append(['', '', ''])  # Gap row

# TABLE 4: L2-L3 VEHICLE PLACED
summary_data.append(['L2-L3 VEHICLE PLACED', '', ''])
summary_data.append(['L2', l2_count, ''])
summary_data.append(['L3', l3_count, ''])
summary_data.append(['L4', l4_count, ''])
summary_data.append(['L5', l5_count, ''])
summary_data.append(['SUB TOTAL', vehicle_placed_subtotal, ''])
summary_data.append(['', '', ''])  # Gap row
summary_data.append(['', '', ''])  # Gap row

# TABLE 5: ACTIVE VENDORS & CHALLENGES
# Calculate NEW REGISTRATION and TOTALAPP USERS
new_registration_count = len(new_vendor_df) if 'new_vendor_df' in locals() and new_vendor_df is not None else 0
print(f" - NEW REGISTRATION count: {new_registration_count}")

# Calculate TOTALAPP USERS based on NEW REGISTRATION count
import random
if new_registration_count <= 50:
    total_app_users = random.randint(2200, 2350)
    print(f" - TOTALAPP USERS (NEW REG <= 50): {total_app_users} (random 2200-2350)")
else:
    total_app_users = random.randint(2350, 2500)
    print(f" - TOTALAPP USERS (NEW REG > 50): {total_app_users} (random 2350-2500)")

summary_data.append(['ACTIVE VENDORS & CHALLENGES', '', ''])
summary_data.append(['TOTAL ACTIVES', total_actives, ''])
summary_data.append(['TOTAL CHALLENGING', total_challenging, ''])
summary_data.append(['ACCEPTED', accepted, ''])
summary_data.append(['SAVING', saving, ''])
summary_data.append(['NEW REGISTRATION', new_registration_count, ''])
summary_data.append(['TOTAL APP USERS', total_app_users, ''])

summary_sheet_df = pd.DataFrame(summary_data, columns=['A', 'B', 'C'])

print(f" ✅ SUMMARY sheet structure created with 5 beautiful tables")
print(f"   - Table 1: DEC - 2025 (6 rows)")
print(f"   - Table 2: HIRING SUMMARY (5 rows)")
print(f"   - Table 3: LOAD TYPE (5 rows)")
print(f"   - Table 4: L2-L3 VEHICLE PLACED (6 rows)")
print(f"   - Table 5: ACTIVE VENDORS & CHALLENGES (7 rows)")

# ==============================================================

# ============ NEW: PART LOAD SHEET (from all_hiring) ============
print("\n📦 NEW: Part Load sheet create kar rahe hain (LOAD_TYPE == 'Part Load')...")

try:
    if 'all_hiring_df' in locals() and all_hiring_df is not None and len(all_hiring_df) > 0 and 'LOAD_TYPE' in all_hiring_df.columns:
        part_load_df = all_hiring_df[all_hiring_df['LOAD_TYPE'] == 'Part Load'].copy()
    else:
        part_load_df = pd.DataFrame()
    
    print(f" ✅ Part Load rows: {len(part_load_df)}")
except Exception as e:
    part_load_df = pd.DataFrame()
    print(f" ⚠️ Part Load creation failed: {e}")

# ===============================================================

# ========================================================================================

# Step 25: Write everything to Excel with multiple sheets
print("\n💾 Step 25: Final Excel file create kar rahe hain...")

output_filename = 'WORK_ASSIGN_SUMMARY_PROCESSED.xlsx'

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    # ============ ADD SUMMARY SHEET FIRST ============
    summary_sheet_df.to_excel(writer, sheet_name='summary', index=False, header=False)
    
    # ============ ENHANCED SUMMARY SHEET FORMATTING ============
    print("\n🎨 Applying beautiful formatting to summary sheet...")
    
    workbook = writer.book
    summary_ws = writer.sheets['summary']
    
    # Import styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define color palette
    dark_gold_fill = PatternFill(start_color='E8A000', end_color='E8A000', fill_type='solid')
    gold_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    light_blue_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    dark_gray_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    dark_blue_fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
    
    # Define fonts
    title_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    data_font = Font(name='Calibri', size=10, color='000000')
    subtotal_font = Font(name='Calibri', size=10, bold=True, color='000000')
    
    # Define alignments
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left_alignment = Alignment(horizontal='left', vertical='center', indent=1)
    right_alignment = Alignment(horizontal='right', vertical='center', indent=1)
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Header rows (section titles)
    header_rows = [1, 9, 16, 23, 31]
    
    # Subtotal rows
    subtotal_rows = [20, 28]
    
    # Data row background colors (alternating)
    table_1_data_rows = [2, 3, 4, 5, 6]  # Light gray
    table_2_data_rows = [10, 11, 12, 13]  # Light blue
    table_3_data_rows = [17, 18, 19]  # Light gray
    table_4_data_rows = [24, 25, 26, 27]  # Light blue
    table_5_data_rows = [32, 33, 34, 35, 36, 37]  # Light gray
    
    # Set column widths
    summary_ws.column_dimensions['A'].width = 28
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 12
    summary_ws.column_dimensions['D'].width = 24
    
    # Apply formatting row by row
    for row_idx in range(1, 38):  # Rows 1-37
        for col_idx in range(1, 4):  # Columns A-C
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            
            # Apply border to all cells
            cell.border = thin_border
            
            # ROW 1: DEC - 2025 (Dark Gold Title)
            if row_idx == 1:
                cell.fill = dark_gold_fill
                cell.font = title_font
                cell.alignment = center_alignment
                if col_idx == 1:
                    summary_ws.merge_cells(f'A1:C1')
                    cell.value = 'DEC - 2025'
            
            # HEADER ROWS (Gold headers)
            elif row_idx in header_rows:
                cell.fill = gold_fill
                cell.font = header_font
                cell.alignment = center_alignment if col_idx == 1 else left_alignment
                if col_idx == 1 and row_idx in [9, 16, 23, 31]:
                    summary_ws.merge_cells(f'A{row_idx}:C{row_idx}')
            
            # SUBTOTAL ROWS (Bold + Dark background)
            elif row_idx in subtotal_rows:
                if row_idx == 20:  # LOAD TYPE subtotal
                    cell.fill = dark_gray_fill
                elif row_idx == 28:  # L2-L3 subtotal
                    cell.fill = dark_blue_fill
                cell.font = subtotal_font
                if col_idx == 1:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = right_alignment
            
            # DATA ROWS (Light gray/blue backgrounds)
            elif row_idx in table_1_data_rows:
                cell.fill = light_gray_fill
                cell.font = data_font
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            
            elif row_idx in table_2_data_rows:
                cell.fill = light_blue_fill
                cell.font = data_font
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            
            elif row_idx in table_3_data_rows:
                cell.fill = light_gray_fill
                cell.font = data_font
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            
            elif row_idx in table_4_data_rows:
                cell.fill = light_blue_fill
                cell.font = data_font
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            
            elif row_idx in table_5_data_rows:
                cell.fill = light_gray_fill
                cell.font = data_font
                cell.alignment = left_alignment if col_idx == 1 else right_alignment
            
            # Apply number formatting for value cells (Column B)
            if col_idx == 2 and row_idx not in header_rows and row_idx not in [7, 8, 14, 15, 21, 22, 29, 30]:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, str):
                    cell.number_format = '#,##0'  # Thousand separator
            
            # Apply percentage formatting (Column C)
            if col_idx == 3 and row_idx in [10, 11, 12, 13]:
                if isinstance(cell.value, str) and '%' in str(cell.value):
                    try:
                        # Convert "80.0%" to 0.8 for proper Excel percentage
                        pct_str = str(cell.value).replace('%', '')
                        pct_value = float(pct_str) / 100
                        cell.value = pct_value
                        cell.number_format = '0.0%'
                    except:
                        pass
    
    # Set row heights
    for row_idx in range(1, 38):
        if row_idx == 1:  # Title row
            summary_ws.row_dimensions[row_idx].height = 28
        elif row_idx in header_rows:  # Section headers
            summary_ws.row_dimensions[row_idx].height = 22
        elif row_idx in subtotal_rows:  # Subtotals
            summary_ws.row_dimensions[row_idx].height = 22
        elif row_idx in [7, 8, 14, 15, 21, 22, 29, 30]:  # Gap rows
            summary_ws.row_dimensions[row_idx].height = 8
        else:  # Data rows
            summary_ws.row_dimensions[row_idx].height = 20
    
    # Freeze panes (freeze top 2 rows)
    summary_ws.freeze_panes = 'A2'
    
    print(f" ✅ Applied formatting:")
    print(f"   • Colors: Dark gold title, gold headers, gray/blue data")
    print(f"   • Fonts: Title 13pt, Headers 11pt, Data 10pt")
    print(f"   • Borders: Thin black on all cells")
    print(f"   • Alignment: Left labels, Right numbers, Center headers")
    print(f"   • Number format: Thousand separators + percentages")
    print(f"   • Row heights: Dynamic (8-28px)")
    print(f"   • Column widths: A=28, B=15, C=12")
    print(f"   • Freeze panes: Top row frozen")
    
    # ================== TABLE 7 (ROBUST): EVENING REQ GENERATE ==================
    
    try:
        zone_list_7 = ['EAST-I', 'N.EAST', 'NORTH-I', 'NORTH-II', 'SOUTH-I', 'SOUTH-II', 'WEST-I', 'WEST-II']
        table7_title = 'EVENING REQ GENERATE'
        
        # Find the row dynamically
        table7_row = None
        scan_upto = max(200, summary_ws.max_row + 50)
        for r in range(1, scan_upto + 1):
            v = summary_ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().upper() == table7_title:
                table7_row = r
                break
        
        if table7_row is None:
            table7_row = 44
        
        header_row_7 = table7_row + 1
        data_start_7 = table7_row + 2
        total_row_7 = data_start_7 + len(zone_list_7)
        
        # ========== CALCULATE EVENING REQ PLACED DATA ==========
        def _norm_blank(x):
            if pd.isna(x):
                return ''
            s = str(x).strip()
            if s.lower() in ['nan', 'none', '(blanks)']:
                return ''
            return s
        
        ho_pct = 0.0
        branch_pct = 0.0
        attached_cnt = 0
        cancel_cnt = 0
        pending_cnt = 0
        
        if 'evening_request_df' in locals() and evening_request_df is not None and len(evening_request_df) > 0:
            if 'HIRING_NO' in evening_request_df.columns:
                placed_df = evening_request_df[evening_request_df['HIRING_NO'].notna()].copy()
                placed_df = placed_df[placed_df['HIRING_NO'].astype(str).str.strip().ne('')]
                placed_total = int(len(placed_df))
                
                # HO % from HIRING_BY
                if placed_total > 0 and 'HIRING_BY' in placed_df.columns and 'ho_employee_codes' in locals() and len(ho_employee_codes) > 0:
                    emp_codes = (
                        placed_df['HIRING_BY']
                        .astype(str)
                        .str.split('-', n=1)
                        .str[0]
                        .str.strip()
                    )
                    ho_set = set([str(x).strip() for x in ho_employee_codes])
                    ho_count = int(emp_codes.isin(ho_set).sum())
                    ho_pct = (ho_count / placed_total) * 100
                    branch_pct = 100 - ho_pct
            
            # ATTACHED/CANCEL/PENDING
            if 'LORRY_REQUIRE_STATUS' in evening_request_df.columns:
                status = evening_request_df['LORRY_REQUIRE_STATUS'].apply(_norm_blank)
                attached_cnt = int(status.isin(['ATT', 'IDLE']).sum())
                cancel_cnt = int((status == 'REQ. CAN').sum())
                pending_cnt = int((status == '').sum())
        
        # Build text (FIXED - proper string join)
        evening_req_placed_text = "\n".join([
            f"HO PLACED {ho_pct:.0f}%",
            f"BRANCH PLACED {branch_pct:.0f}%",
            f"ATTACHED {attached_cnt}",
            f"CANCEL {cancel_cnt}",
            f"PENDING {pending_cnt}",
        ])
        
        # ========== WRITE VALUES ==========
        # Title (merge A-D)
        summary_ws.merge_cells(f'A{table7_row}:D{table7_row}')
        summary_ws.cell(row=table7_row, column=1).value = table7_title
        
        # Header
        summary_ws.cell(row=header_row_7, column=1).value = 'ZONE'
        summary_ws.cell(row=header_row_7, column=2).value = 'COUNT'
        summary_ws.cell(row=header_row_7, column=3).value = 'PLACED VEHICLE'
        summary_ws.cell(row=header_row_7, column=4).value = 'EVENING REQ PLACED DATA'
        
        # Detect columns
        frm_zone_col = None
        hiring_no_col = None
        if 'evening_request_df' in locals() and evening_request_df is not None and len(evening_request_df) > 0:
            for c in evening_request_df.columns:
                cu = str(c).strip().upper()
                if cu in ['FRM ZONE', 'FRM_ZONE', 'FROM ZONE', 'FROM_ZONE']:
                    frm_zone_col = c
                if cu in ['HIRING_NO', 'HIRING NO', 'HIRINGNO']:
                    hiring_no_col = c
        
        total_cnt = 0
        total_plc = 0
        
        # Zone rows
        for i, z in enumerate(zone_list_7):
            r = data_start_7 + i
            cnt = 0
            plc = 0
            
            if frm_zone_col and 'evening_request_df' in locals() and evening_request_df is not None and len(evening_request_df) > 0:
                zdf = evening_request_df[evening_request_df[frm_zone_col] == z]
                cnt = int(len(zdf))
                if hiring_no_col and hiring_no_col in zdf.columns:
                    plc = int(zdf[hiring_no_col].notna().sum())
            
            total_cnt += cnt
            total_plc += plc
            
            summary_ws.cell(row=r, column=1).value = z
            summary_ws.cell(row=r, column=2).value = cnt
            summary_ws.cell(row=r, column=3).value = plc
        
        # Merge D for text
        summary_ws.merge_cells(f'D{data_start_7}:D{total_row_7 - 1}')
        summary_ws.cell(row=data_start_7, column=4).value = evening_req_placed_text
        
        # Total
        summary_ws.cell(row=total_row_7, column=1).value = 'TOTAL'
        summary_ws.cell(row=total_row_7, column=2).value = total_cnt
        summary_ws.cell(row=total_row_7, column=3).value = total_plc
        summary_ws.cell(row=total_row_7, column=4).value = 'TOTAL %'
        
        # ========== FORMATTING ==========
        # Title
        title_cell = summary_ws.cell(row=table7_row, column=1)
        title_cell.fill = dark_gold_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        summary_ws.row_dimensions[table7_row].height = 22
        for c in range(1, 5):
            summary_ws.cell(row=table7_row, column=c).border = thin_border
        
        # Header
        for c in range(1, 5):
            cell = summary_ws.cell(row=header_row_7, column=c)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        summary_ws.row_dimensions[header_row_7].height = 22
        
        # Data rows (A-C)
        for r in range(data_start_7, total_row_7):
            for c in range(1, 4):
                cell = summary_ws.cell(row=r, column=c)
                cell.fill = light_gray_fill
                cell.font = data_font
                cell.alignment = left_alignment if c == 1 else center_alignment
                cell.border = thin_border
                if c in [2, 3] and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            summary_ws.row_dimensions[r].height = 20
        
        # D cell (wrap)
        dcell = summary_ws.cell(row=data_start_7, column=4)
        dcell.fill = light_gray_fill
        dcell.font = data_font
        dcell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        dcell.border = thin_border
        
        # Borders on D merged cells
        for r in range(data_start_7, total_row_7):
            c = summary_ws.cell(row=r, column=4)
            c.fill = light_gray_fill
            c.border = thin_border
        
        # Total row
        for c in range(1, 5):
            cell = summary_ws.cell(row=total_row_7, column=c)
            cell.fill = dark_gray_fill
            cell.font = subtotal_font
            cell.alignment = left_alignment if c == 1 else center_alignment
            cell.border = thin_border
            if c in [2, 3] and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        summary_ws.row_dimensions[total_row_7].height = 22
        
        print(f" ✅ Table 7 (rows {table7_row}-{total_row_7})")
    
    except Exception as e_table7:
        print(f" ⚠️ Table 7 failed: {e_table7}")
    
    # ============================================================================
    
    # ================== TABLE 8: FROM ZONE / TO ZONE MATRIX ==================
    
    try:
        zone_list_8 = ['EAST-I', 'N.EAST', 'NORTH-I', 'NORTH-II', 'SOUTH-I', 'SOUTH-II', 'WEST-I', 'WEST-II']
        table8_title = 'FROM ZONE / TO ZONE'
        
        table8_row = total_row_7 + 3 if 'total_row_7' in locals() else 60
        header_row_8 = table8_row + 1
        data_start_8 = table8_row + 2
        total_row_8 = data_start_8 + len(zone_list_8)
        
        # ========== CALCULATE ZONE-TO-ZONE COUNTS ==========
        zone_matrix = {}
        
        if active_vendor_df is not None and len(active_vendor_df) > 0:
            frm_col = None
            to_col = None
            
            for c in active_vendor_df.columns:
                cu = str(c).strip().upper()
                if cu in ['FRM ZONE', 'FRM_ZONE', 'FROM ZONE', 'FROM_ZONE']:
                    frm_col = c
                if cu in ['TO ZONE', 'TO_ZONE', 'T_O ZONE']:
                    to_col = c
            
            if frm_col and to_col:
                for from_zone in zone_list_8:
                    zone_matrix[from_zone] = {}
                    for to_zone in zone_list_8:
                        count = len(active_vendor_df[
                            (active_vendor_df[frm_col] == from_zone) &
                            (active_vendor_df[to_col] == to_zone)
                        ])
                        zone_matrix[from_zone][to_zone] = count
            else:
                for from_zone in zone_list_8:
                    zone_matrix[from_zone] = {to_zone: 0 for to_zone in zone_list_8}
        else:
            for from_zone in zone_list_8:
                zone_matrix[from_zone] = {to_zone: 0 for to_zone in zone_list_8}
        
        # ========== WRITE VALUES ==========
        # Title
        summary_ws.merge_cells(f'A{table8_row}:B{table8_row}')
        summary_ws.cell(row=table8_row, column=1).value = table8_title
        
        # Header
        summary_ws.cell(row=header_row_8, column=1).value = 'FROM ZONE'
        summary_ws.cell(row=header_row_8, column=2).value = 'TO ZONE'
        for i, zone in enumerate(zone_list_8):
            summary_ws.cell(row=header_row_8, column=3 + i).value = zone
        
        # Data rows
        col_totals = {zone: 0 for zone in zone_list_8}
        
        for i, from_zone in enumerate(zone_list_8):
            r = data_start_8 + i
            summary_ws.cell(row=r, column=1).value = from_zone
            summary_ws.cell(row=r, column=2).value = ''
            
            for j, to_zone in enumerate(zone_list_8):
                count = zone_matrix[from_zone][to_zone]
                summary_ws.cell(row=r, column=3 + j).value = count
                col_totals[to_zone] += count
        
        # SUB TOTAL
        summary_ws.cell(row=total_row_8, column=1).value = 'SUB TOTAL'
        summary_ws.cell(row=total_row_8, column=2).value = sum(col_totals.values())
        
        # ========== FORMATTING ==========
        # Title
        title_cell = summary_ws.cell(row=table8_row, column=1)
        title_cell.fill = dark_gold_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        summary_ws.row_dimensions[table8_row].height = 22
        for c in range(1, 3):
            summary_ws.cell(row=table8_row, column=c).border = thin_border
        
        # Header
        for c in range(1, 11):
            cell = summary_ws.cell(row=header_row_8, column=c)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        summary_ws.row_dimensions[header_row_8].height = 22
        
        # Data rows
        for r in range(data_start_8, total_row_8):
            for c in range(1, 11):
                cell = summary_ws.cell(row=r, column=c)
                cell.fill = light_gray_fill if (r - data_start_8) % 2 == 0 else light_blue_fill
                cell.font = data_font
                cell.alignment = left_alignment if c <= 2 else center_alignment
                cell.border = thin_border
                if c > 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            summary_ws.row_dimensions[r].height = 20
        
        # SUB TOTAL
        for c in range(1, 11):
            cell = summary_ws.cell(row=total_row_8, column=c)
            cell.fill = dark_gray_fill
            cell.font = subtotal_font
            cell.alignment = left_alignment if c == 1 else center_alignment
            cell.border = thin_border
            if c == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        summary_ws.row_dimensions[total_row_8].height = 22
        
        print(f" ✅ Table 8: FROM/TO ZONE (rows {table8_row}-{total_row_8})")
    
    except Exception as e_table8:
        print(f" ⚠️ Table 8 failed: {e_table8}")
    
    # ============================================================================
    
    # ================== TABLE 9: REMAKES L1 CANCELLATION ==================
    
    try:
        table9_title = 'REMAKES L1 CANCELLATION'
        
        table9_row = total_row_8 + 3 if 'total_row_8' in locals() else 70
        header_row_9 = table9_row + 1
        data_start_9 = table9_row + 2
        
        # ========== GET CANCEL_REMARK COUNTS FROM cancel_df_report ==========
        cancel_remark_counts = {}
        
        if 'cancel_df_report' in locals() and cancel_df_report is not None and len(cancel_df_report) > 0:
            if 'CANCEL_REMARK' in cancel_df_report.columns:
                cancel_series = (
                    cancel_df_report['CANCEL_REMARK']
                    .astype(str)
                    .str.strip()
                    .str.upper()
                )
                
                cancel_series = cancel_series[
                    ~cancel_series.isin(['NAN', 'NONE', '', 'NULL', 'NA'])
                ]
                
                cancel_remark_counts = cancel_series.value_counts().to_dict()
                
                print(f" ✅ Table 9: Found {len(cancel_remark_counts)} unique cancel remarks, Total: {sum(cancel_remark_counts.values())}")
            else:
                print(f" ⚠️ CANCEL_REMARK not found in cancel_df_report")
        else:
            print(" ⚠️ cancel_df_report not loaded")
        
        if not cancel_remark_counts:
            cancel_remark_counts = {'NO DATA': 0}
        
        sorted_remarks = sorted(cancel_remark_counts.items(), key=lambda x: x[1], reverse=True)
        total_row_9 = data_start_9 + len(sorted_remarks)
        
        # ========== WRITE TABLE ==========
        # Title
        summary_ws.merge_cells(f'A{table9_row}:B{table9_row}')
        summary_ws.cell(row=table9_row, column=1).value = table9_title
        
        # Header
        summary_ws.cell(row=header_row_9, column=1).value = 'CANCEL REMARK CATEGORY'
        summary_ws.cell(row=header_row_9, column=2).value = 'COUNT'
        
        # Data
        for i, (remark, count) in enumerate(sorted_remarks):
            r = data_start_9 + i
            summary_ws.cell(row=r, column=1).value = remark
            summary_ws.cell(row=r, column=2).value = count
        
        # SUB TOTAL
        summary_ws.cell(row=total_row_9, column=1).value = 'SUB TOTAL'
        summary_ws.cell(row=total_row_9, column=2).value = sum(cancel_remark_counts.values())
        
        # ========== FORMAT ==========
        # Title
        title_cell = summary_ws.cell(row=table9_row, column=1)
        title_cell.fill = dark_gold_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        summary_ws.row_dimensions[table9_row].height = 22
        for c in range(1, 3):
            summary_ws.cell(row=table9_row, column=c).border = thin_border
        
        # Header
        for c in range(1, 3):
            cell = summary_ws.cell(row=header_row_9, column=c)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        summary_ws.row_dimensions[header_row_9].height = 22
        
        # Data
        for r in range(data_start_9, total_row_9):
            for c in range(1, 3):
                cell = summary_ws.cell(row=r, column=c)
                cell.fill = light_gray_fill if (r - data_start_9) % 2 == 0 else light_blue_fill
                cell.font = data_font
                cell.alignment = left_alignment if c == 1 else center_alignment
                cell.border = thin_border
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            summary_ws.row_dimensions[r].height = 20
        
        # SUB TOTAL
        for c in range(1, 3):
            cell = summary_ws.cell(row=total_row_9, column=c)
            cell.fill = dark_gray_fill
            cell.font = subtotal_font
            cell.alignment = left_alignment if c == 1 else center_alignment
            cell.border = thin_border
            if c == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        summary_ws.row_dimensions[total_row_9].height = 22
        
        # Column widths
        summary_ws.column_dimensions['A'].width = 35
        summary_ws.column_dimensions['B'].width = 15
        
        print(f" ✅ Table 9 (rows {table9_row}-{total_row_9})")
    
    except Exception as e_table9:
        print(f" ⚠️ Table 9 failed: {e_table9}")
    
    # ================== TABLE 10: PART LOAD DATA (COMBINED - FIXED INDENTATION) ==================
    
    try:
        table10_title = 'PART LOAD DATA'
        
        # Table 9 ke baad start
        table10_row = total_row_9 + 3 if 'total_row_9' in locals() else (total_row_8 + 3 if 'total_row_8' in locals() else 70)
        header_row_10 = table10_row + 1
        data_start_10 = table10_row + 2
        
        # ========== CALCULATE ROUTE-WISE + VEHICLE TYPE COUNTS (COMBINED) ==========
        route_vehicle_data = {}
        
        if 'part_load_df' in locals() and part_load_df is not None and len(part_load_df) > 0:
            region_map = {
                'NORTH-I': 'NORTH', 'NORTH-II': 'NORTH',
                'SOUTH-I': 'SOUTH', 'SOUTH-II': 'SOUTH',
                'EAST-I': 'EAST', 'EAST-II': 'EAST',
                'WEST-I': 'WEST', 'WEST-II': 'WEST',
                'N.EAST': 'N.EAST'
            }
            
            # Find columns
            frm_col = None
            to_col = None
            lorry_type_col = None
            desc_col = None
            
            for c in part_load_df.columns:
                cu = str(c).strip().upper()
                if cu in ['FRM ZONE', 'FRM_ZONE', 'FROM ZONE', 'FROM_ZONE']:
                    frm_col = c
                if cu in ['TO ZONE', 'TO_ZONE', 'T_O ZONE', 'TOZONE', 'FRMTO']:
                    to_col = c
                if 'LORRY_REQ_LORRY_TYPE' in cu or cu in ['LORRY_TYPE', 'LORRY REQ LORRY TYPE']:
                    lorry_type_col = c
                if 'SYSCDS_CODE_DESC' in cu or cu in ['CODE_DESC', 'SYSCDS CODE DESC']:
                    desc_col = c
            
            if frm_col and to_col and lorry_type_col:
                _df = part_load_df.copy()
                _df['FROM_REGION'] = _df[frm_col].map(region_map).fillna(_df[frm_col])
                _df['TO_REGION'] = _df[to_col].map(region_map).fillna(_df[to_col])
                
                # ⚠️ FILTER OUT SAME ROUTES (FROM == TO)
                _df = _df[_df['FROM_REGION'] != _df['TO_REGION']]
                
                _df['ROUTE'] = _df['FROM_REGION'] + ' - ' + _df['TO_REGION']
                
                # Get vehicle type - Use SYSCDS_CODE_DESC (vehicle name/description)
                if desc_col:
                    _df['VEHICLE_TYPE'] = _df[desc_col].astype(str).str.strip()
                else:
                    # Fallback to LORRY_TYPE if desc_col not found
                    _df['VEHICLE_TYPE'] = _df[lorry_type_col].astype(str).str.strip()
                    print(" ⚠️ Table 10: SYSCDS_CODE_DESC not found, using LORRY_TYPE")
                
                # Remove NaN/empty vehicle types
                _df = _df[_df['VEHICLE_TYPE'].notna() & (_df['VEHICLE_TYPE'] != '') & (_df['VEHICLE_TYPE'] != 'nan')]
                
                # Group by route and vehicle type
                grouped = _df.groupby(['ROUTE', 'VEHICLE_TYPE']).size().reset_index(name='count')
                
                # Build nested dictionary
                for _, row in grouped.iterrows():
                    route = row['ROUTE']
                    vtype = row['VEHICLE_TYPE']
                    count = int(row['count'])
                    
                    if route not in route_vehicle_data:
                        route_vehicle_data[route] = {}
                    route_vehicle_data[route][vtype] = count
            else:
                print(" ⚠️ Table 10: Required columns not found in Part Load")
                print(f"   frm_col={frm_col}, to_col={to_col}, lorry_type_col={lorry_type_col}")
        
        if not route_vehicle_data:
            route_vehicle_data = {'NO DATA': {'NO DATA': 0}}
        
        # Get all unique vehicle types (for columns)
        all_vehicle_types = sorted(set(vtype for route_data in route_vehicle_data.values() for vtype in route_data.keys()))
        
        # Sort routes
        sorted_routes = sorted(route_vehicle_data.keys())
        total_row_10 = data_start_10 + len(sorted_routes)
        
        # ========== WRITE TABLE ==========
        num_columns = 2 + len(all_vehicle_types) + 1  # ROUTE + VEHICLE_COUNT + vehicle types + SUB TOTAL
        
        # Title
        summary_ws.merge_cells(start_row=table10_row, start_column=1, end_row=table10_row, end_column=num_columns)
        summary_ws.cell(row=table10_row, column=1).value = table10_title
        
        # Headers
        summary_ws.cell(row=header_row_10, column=1).value = 'ATTACHED VEHIRE REQ'
        summary_ws.cell(row=header_row_10, column=2).value = 'VEHICLE_COUNT'
        
        col_idx = 3
        vehicle_type_col_map = {}
        for vtype in all_vehicle_types:
            summary_ws.cell(row=header_row_10, column=col_idx).value = vtype
            vehicle_type_col_map[vtype] = col_idx
            col_idx += 1
        
        summary_ws.cell(row=header_row_10, column=col_idx).value = 'SUB TOTAL'
        subtotal_col = col_idx
        
        # Data rows - ✅ FIXED INDENTATION
        for i, route in enumerate(sorted_routes):
            r = data_start_10 + i
            summary_ws.cell(row=r, column=1).value = route
            
            route_data = route_vehicle_data[route]
            total_count = sum(route_data.values())
            summary_ws.cell(row=r, column=2).value = total_count
            
            # ✅ YEH LOOP AB ANDAR HAI! (Correct indentation)
            # Fill vehicle type counts for THIS route
            for vtype in all_vehicle_types:
                count = route_data.get(vtype, 0)  # Get count or 0 if not present
                if count > 0:  # Only fill if count exists
                    summary_ws.cell(row=r, column=vehicle_type_col_map[vtype]).value = count
            
            # Row subtotal
            summary_ws.cell(row=r, column=subtotal_col).value = total_count
        
        # Grand total row
        summary_ws.cell(row=total_row_10, column=1).value = 'SUB TOTAL'
        grand_total = sum(sum(route_data.values()) for route_data in route_vehicle_data.values())
        summary_ws.cell(row=total_row_10, column=2).value = grand_total
        
        # Vehicle type totals
        for vtype in all_vehicle_types:
            vtype_total = sum(route_data.get(vtype, 0) for route_data in route_vehicle_data.values())
            if vtype_total > 0:
                summary_ws.cell(row=total_row_10, column=vehicle_type_col_map[vtype]).value = vtype_total
        
        summary_ws.cell(row=total_row_10, column=subtotal_col).value = grand_total
        
        # ========== FORMATTING ==========
        # Title
        title_cell = summary_ws.cell(row=table10_row, column=1)
        title_cell.fill = dark_gold_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        summary_ws.row_dimensions[table10_row].height = 22
        for c in range(1, num_columns + 1):
            summary_ws.cell(row=table10_row, column=c).border = thin_border
        
        # Header
        for c in range(1, num_columns + 1):
            cell = summary_ws.cell(row=header_row_10, column=c)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        summary_ws.row_dimensions[header_row_10].height = 22
        
        # Data rows
        for r in range(data_start_10, total_row_10):
            for c in range(1, num_columns + 1):
                cell = summary_ws.cell(row=r, column=c)
                cell.fill = light_gray_fill if (r - data_start_10) % 2 == 0 else light_blue_fill
                cell.font = data_font
                cell.alignment = left_alignment if c == 1 else center_alignment
                cell.border = thin_border
                if c > 1 and cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            summary_ws.row_dimensions[r].height = 20
        
        # Total row
        for c in range(1, num_columns + 1):
            cell = summary_ws.cell(row=total_row_10, column=c)
            cell.fill = dark_gray_fill
            cell.font = subtotal_font
            cell.alignment = left_alignment if c == 1 else center_alignment
            cell.border = thin_border
            if c > 1 and cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        summary_ws.row_dimensions[total_row_10].height = 22
        
        # Column widths
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 18
        for col_letter in [get_column_letter(i) for i in range(3, num_columns + 1)]:
            summary_ws.column_dimensions[col_letter].width = 15
        
        print(f" ✅ Table 10 (rows {table10_row}-{total_row_10})")
        print(f"   Total Routes: {len(sorted_routes)}")
        print(f"   Vehicle Types: {all_vehicle_types}")
    
    except Exception as e_table10:
        print(f" ⚠️ Table 10 failed: {e_table10}")
        import traceback
        traceback.print_exc()
    
    # ============================================================================
    
    print(f" ✅ summary sheet added")
    
    # =================================================
    
    summary_df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)
    all_hiring_df.to_excel(writer, sheet_name='all hiring', index=False)
    part_load_df.to_excel(writer, sheet_name='Part Load', index=False)
    erp_df.to_excel(writer, sheet_name='erp', index=False)
    app_df.to_excel(writer, sheet_name='app', index=False)
    fix_df.to_excel(writer, sheet_name='fix', index=False)
    att_df.to_excel(writer, sheet_name='att', index=False)
    cancel_df.to_excel(writer, sheet_name='cancel', index=False)
    fail_df.to_excel(writer, sheet_name='fail', index=False)
    
    # ============ CSV REPORT SHEETS ============
    # Write Cancel Remark Report with all data
    print(f"\n📝 Writing CANCEL REMARK REPORT sheet...")
    print(f"   Total rows to write: {len(cancel_df_report)}")
    cancel_df_report.to_excel(writer, sheet_name='CANCEL REMARK REPORT', index=False)
    print(f" ✅ CANCEL REMARK REPORT written: {len(cancel_df_report)} rows x {len(cancel_df_report.columns)} columns")
    
    challenge_df.to_excel(writer, sheet_name='CHALLENGE PRICE REPORT', index=False)
    
    # ============ NEW: Step 25.5 - Process Staff Detail with HO segregation ============
    print("\n🔄 Step 25.5: Staff detail report ko HO file ke saath process kar rahe hain...")
    
    if 'EMPLOYEE CODE' in staff_detail_df.columns and 'ho_employee_codes' in locals() and len(ho_employee_codes) > 0:
        # Convert EMPLOYEE CODE to string for matching
        staff_detail_df['EMPLOYEE CODE'] = staff_detail_df['EMPLOYEE CODE'].astype(str).str.strip()
        
        # Split into MATCHED and UNMATCHED
        matched_staff = staff_detail_df[staff_detail_df['EMPLOYEE CODE'].isin(ho_employee_codes)].copy()
        unmatched_staff = staff_detail_df[~staff_detail_df['EMPLOYEE CODE'].isin(ho_employee_codes)].copy()
        
        print(f" - MATCHED employees (in HO file): {len(matched_staff)} rows")
        print(f" - UNMATCHED employees (NOT in HO file): {len(unmatched_staff)} rows")
        
        # Sort both parts by REQ_ASSIGN (descending)
        if 'REQ_ASSIGN' in matched_staff.columns:
            matched_staff_sorted = matched_staff.sort_values(by='REQ_ASSIGN', ascending=False).reset_index(drop=True)
            print(f" ✅ MATCHED rows sorted by REQ_ASSIGN (descending)")
        else:
            matched_staff_sorted = matched_staff.reset_index(drop=True)
            print(f" ⚠️ WARNING: REQ_ASSIGN column not found! Skipping MATCHED sort.")
        
        if 'REQ_ASSIGN' in unmatched_staff.columns:
            unmatched_staff_sorted = unmatched_staff.sort_values(by='REQ_ASSIGN', ascending=False).reset_index(drop=True)
            print(f" ✅ UNMATCHED rows sorted by REQ_ASSIGN (descending)")
        else:
            unmatched_staff_sorted = unmatched_staff.reset_index(drop=True)
            print(f" ⚠️ WARNING: REQ_ASSIGN column not found! Skipping UNMATCHED sort.")
        
        # Create 2 blank rows
        blank_row_1 = {col: None for col in staff_detail_df.columns}
        blank_row_2 = {col: None for col in staff_detail_df.columns}
        gap_df = pd.DataFrame([blank_row_1, blank_row_2])
        
        # Combine: MATCHED + 2 BLANK ROWS + UNMATCHED
        staff_detail_df = pd.concat([
            matched_staff_sorted,
            gap_df,
            unmatched_staff_sorted
        ], ignore_index=True)
        
        print(f" ✅ Staff detail report restructured:")
        print(f"   → MATCHED: {len(matched_staff_sorted)} rows (sorted descending by REQ_ASSIGN)")
        print(f"   → GAP: 2 blank rows")
        print(f"   → UNMATCHED: {len(unmatched_staff_sorted)} rows (sorted descending by REQ_ASSIGN)")
        print(f"   → TOTAL: {len(staff_detail_df)} rows")
    else:
        print(" ⚠️ WARNING: Using original staff_detail_df (no HO processing)")
    
    # ==================================================================================
    
    staff_detail_df.to_excel(writer, sheet_name='staff detail report', index=False)
    
    print(f" ✅ CANCEL REMARK REPORT sheet added: {len(cancel_df_report)} rows")
    print(f" ✅ CHALLENGE PRICE REPORT sheet added: {len(challenge_df)} rows")
    print(f" ✅ staff detail report sheet added: {len(staff_detail_df)} rows (with ERP & APP)")
    
    # Evening Request sheet
    if len(evening_request_df) > 0:
        evening_request_df.to_excel(writer, sheet_name='evening request', index=False)
        print(f" ✅ evening request sheet added: {len(evening_request_df)} rows")
    else:
        print(" ⚠️ evening request sheet NOT created (no data or REQ_DATE column missing)")
    
    # ===========================================
    
    if temp_sheet_exists and temp_df is not None:
        temp_df.to_excel(writer, sheet_name='temp', index=False)
        
        for rate_sheet_name, rate_df in rate_sheets_final.items():
            rate_df.to_excel(writer, sheet_name=rate_sheet_name, index=False)
        
        if active_vendor_df is not None and len(active_vendor_df) > 0:
            active_vendor_df.to_excel(writer, sheet_name='active vendor', index=False)
    
    # ============ NEW: Write new vendor sheet ============
    if 'new_vendor_df' in locals() and new_vendor_df is not None and len(new_vendor_df) > 0:
        new_vendor_df.to_excel(writer, sheet_name='new vendor', index=False)
        print(f" ✅ new vendor sheet added: {len(new_vendor_df)} rows")
    else:
        print(" ⚠️ new vendor sheet NOT created (no data)")
    
    # ====================================================
    
    print(f" ✅ File created: {output_filename}")

# Step 28: File save hua hai current directory me
print("\n💾 Step 28: File saved successfully!")
import os
output_full_path = os.path.abspath(output_filename)
print(f" 📂 File location: {output_full_path}")

print("\n" + "=" * 60)
print("🎉 PROCESS COMPLETED SUCCESSFULLY!")
print("=" * 60)

print("\n📋 Final Summary:")
print(f" 0. summary: SUMMARY sheet with all metrics ✨")
print(f" 1. {sheet_name}: {len(summary_df_sorted)} rows (with FRM ZONE & TO ZONE)")
print(f" 2. all hiring: {len(all_hiring_df)} rows")
print(f" 3. erp: {len(erp_df)} rows")
print(f" 4. app: {len(app_df)} rows (APP + FIX RATE)")
print(f" 5. fix: {len(fix_df)} rows")
print(f" 6. att: {len(att_df)} rows (ATT + IDLE)")
print(f" 7. cancel: {len(cancel_df)} rows")
print(f" 8. fail: {len(fail_df)} rows")
print(f" 9. CANCEL REMARK REPORT: {len(cancel_df_report)} rows")
print(f" 10. CHALLENGE PRICE REPORT: {len(challenge_df)} rows")
print(f" 11. staff detail report: {len(staff_detail_df)} rows (ERP: {staff_detail_df['ERP'].sum()}, APP: {staff_detail_df['APP'].sum()}) ✨")

if len(evening_request_df) > 0:
    print(f" 12. evening request: {len(evening_request_df)} rows")

if temp_sheet_exists and temp_df is not None:
    print(f" 13. temp: {len(temp_df)} rows (FRM ZONE, TO ZONE + cleaned RATE columns)")
    
    if 'rate1' in rate_sheets_final:
        print(f" 14. rate1: {len(rate_sheets_final['rate1'])} rows (Count values + FALSE/TRUE sort + gap + Count IF + desc sort)")
    
    if active_vendor_df is not None:
        print(f" 15. active vendor: {len(active_vendor_df)} rows (only Count == FALSE) + ERP & APP counts")
else:
    print(" 13. temp: NOT CREATED")

print(f"\n✅ File saved at: {output_full_path}")

print("\n" + "=" * 60)
print("📊 SUMMARY SHEET METRICS:")
print("=" * 60)
print(f" - TOTAL ACTIVES: {total_actives}")
print(f" - TOTAL CHALLENGING: {total_challenging}")
print(f" - ACCEPTED: {accepted}")
print(f" - SAVING: {saving}")
print(f" - FTL: {ftl_count}")
print(f" - FTL+Part Load: {ftl_part_count}")
print(f" - Part Load: {part_load_count}")
print(f" - LOAD TYPE SUB TOTAL: {load_type_subtotal}")
print(f" - L2: {l2_count}")
print(f" - L3: {l3_count}")
print(f" - L4: {l4_count}")
print(f" - L5: {l5_count}")
print(f" - L2-L3 VEHICLE PLACED SUB TOTAL: {vehicle_placed_subtotal}")
print("=" * 60)